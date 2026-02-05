
import openpyxl
import os

file_path = "data/AI_CyberSecurity_Module.xlsx"
sheet_name = "Scoring Matrix"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        exit(1)
        
    ws = wb[sheet_name]
    
    # Read first row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = row
        break
        
    print("Columns found:")
    for col in headers:
        print(f" - {col}")
        
except Exception as e:
    print(f"Error reading excel with openpyxl: {e}")
